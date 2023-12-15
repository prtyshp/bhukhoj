import json
import os
import requests
import sqlite3
from flask import Flask, render_template, request, flash, jsonify, send_file, redirect, url_for
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy.sql import text
from sqlalchemy import and_, desc, or_
from flask_bootstrap import Bootstrap5
from flask_wtf import FlaskForm
from wtforms import SubmitField, SelectField, RadioField, HiddenField, StringField, IntegerField, FloatField
from wtforms.validators import InputRequired, Length, Regexp, NumberRange, Optional
from datetime import date
from basedata import land_types, all_data 
from message_helper import get_templated_message_input, get_text_message_input, send_message, upload_media, get_media_url

import openpyxl
import jinja2
import pdfkit
from heyoo import WhatsApp

# token = 'EAAEFGwYP5GgBO3Rf1BVwYagfUVJpF27ZBixvMQoRMBM6voqzaQRamuvbmoJJ7eSzrYMjuUVYoFmliR1Py3S6nEU64t1oCGaI0got0JMZBLan15RbVo5HxHUTWnHGbknFwndbBW72ZBorgZBrd9CPymeJRU9ZBrQ8UwemDyOXEYKRwNOtTcgIBhVd2ZBCoJ4IeInxHiZCtwMUAFt3EUZD'
# access_token = 'EAAEFGwYP5GgBO3NpVZA14l2R7SquZBr22X2Fw3CYcvxcwQc5YKZA5WEqfLXLZBZBGTkuuUwGem7UEmbk5hZAsOSK9NYIdgWLzmDIcOpi4ZB75qhnHomZA27wke2NRkm5sqh69NJlyT2asmZAprpWTOXh15nFZBRAZBTwyDbNZB6lV0O7s6kH9GPfAq5jTL9tUFfQ'
# phone_number_id = '100140579846345'
# to_phone_number = '+91 81759 89979'
with open('config.json') as config_file:
    config = json.load(config_file)
        
messenger = WhatsApp(token=config['ACCESS_TOKEN'], phone_number_id=config['PHONE_NUMBER_ID'])
# message = 'Hello, world!'
# print("not sent yet")
# messenger.send_message(message=message, recipient_id=to_phone_number)
# print('sent')
app = Flask(__name__)
db_name = "land_records.db"

project_dir = os.path.dirname(os.path.abspath(__file__))
database_file = "sqlite:///{}".format(os.path.join(project_dir, db_name))

app.config["SQLALCHEMY_DATABASE_URI"] = database_file
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = True
app.config['SECRET_KEY'] = "MLXH243GssUWwKdTWS7FDhdwYF56wPj8"
bootstrap = Bootstrap5(app)
db = SQLAlchemy(app)


district_tehsil_village_mapping = all_data
# Data structure for district - tehsil - village mapping
#district_tehsil_village_mapping = {'Hardoi': {'Hardoi': {'140335': 'Akbarpur'}}}
# district_tehsil_village_mapping = {
#     'Hardoi': {
#         'Bilgram': {'140422': 'Adampur', '140751': 'Akbarpur'},
#         'Hardoi': {'140255': 'Abdulpur', '139958': 'Achauli'},
#         'Sandila': {'140946': 'Aant', '140815': 'Adampur'},
#         'Swayjpur': {'139612': 'Adampur', '139616': 'Ahmadpur'},
#         'Shahabad': {'139407': 'Abdulla Nagar', '139226': 'Abdullapur'}
#     }
# }
#tehsil_map = {district: list(data.keys()) for district, data in district_tehsil_village_mapping.items()}    

# Helper function to get unique values from a list of dictionaries
def unique_values(data_list, key):
    return list({item[key]: item for item in data_list}.values())

class LandRecord(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    # Add these three from your data
    district = db.Column(db.String(20))
    tehsil = db.Column(db.String(20))
    village = db.Column(db.String(20))
    
    village_code = db.Column(db.String(20))
    land_type = db.Column(db.String(20))
    khata_number = db.Column(db.String(20))
    fasli_year = db.Column(db.String(20))
    khasra_no = db.Column(db.String(20))
    area = db.Column(db.Float)
    name = db.Column(db.String(100))

class LandSearchForm(FlaskForm):
    district = SelectField('District', validators=[Optional()])
    tehsil = SelectField('Tehsil', validators=[Optional()])
    village = SelectField('Village', validators=[Optional()])
    land_type = SelectField('Land Type', validators=[Optional()])
    area = FloatField('Minimum Area (in acres)', validators=[InputRequired(), NumberRange(min=0.0)])
    submit = SubmitField('Search')

def serialize_land_records(records):
    """
    Helper function to serialize LandRecord objects to a list of dictionaries.
    """
    return [
        {
            'district': record.district,
            'tehsil': record.tehsil,
            'village': record.village,
            'village_code': record.village_code,
            'land_type': record.land_type,
            'khata_number': record.khata_number,
            'fasli_year': record.fasli_year,
            'khasra_no': record.khasra_no,
            'area': record.area,
            'name': record.name,
        }
        for record in records
    ]

def deserialize_land_records(serialized_records):
    """
    Deserialize the list of dictionaries to a list of LandRecord objects.

    Parameters:
        serialized_records (list): List of dictionaries representing the serialized LandRecord objects.

    Returns:
        list: List of LandRecord objects.
    """
    deserialized_records = []
    for record in serialized_records:
        deserialized_record = LandRecord(
            district=record['district'],
            tehsil=record['tehsil'],
            village=record['village'],
            village_code=record['village_code'],
            land_type=record['land_type'],
            khata_number=record['khata_number'],
            fasli_year=record['fasli_year'],
            khasra_no=record['khasra_no'],
            area=record['area'],
            name=record['name']
        )
        deserialized_records.append(deserialized_record)
    return deserialized_records

greeting_flag = bool(False)


def is_database_empty():
    num_records = LandRecord.query.count()
    return num_records == 0

# Initialize the database
# Create the database and tables before the first request
@app.before_request
def create_dbase():
    print(is_database_empty())
    if is_database_empty():
        db.create_all()

        #districts = ['Hardoi']
        #tehsils = { 'Hardoi': ['Bilgram', 'Shahabad', 'Sandila', 'Swayjpur', 'Hardoi'] }

        #api_url_1 = 'https://upbhulekh.gov.in/WS_'
        #api_url_2 = '/service'
        #params = {'village': '140335'}
        # Insert the JSON data into the database
        #api_url = api_url_1 + district + api_url_2
        ap_url = 'https://upbhulekh.gov.in/WS_Hardoi/service'
        for district, tehsil_map in district_tehsil_village_mapping.items():
            print(district)
            for tehsil, village_map in tehsil_map.items():
                print(tehsil)
                for vill_code in village_map.keys():
                    params = {'village': vill_code}
                    #print(vill_code, village_map[vill_code])
                    response = requests.get(ap_url, params=params)
                    data = response.json()  # Parse the JSON data
                    if 'DATA' in data and 'DATA' in data['DATA']:
                        for record in data['DATA']['DATA']:
                            village_code = record['village_code_census']
                            khasra_no = record['khasra_no']
                            # Check if the record already exists in the database
                            existing_record = LandRecord.query.filter_by(village_code=village_code, khasra_no=khasra_no).first()
                            if existing_record is None:
                                new_record = LandRecord(
                                district = district,
                                tehsil = tehsil,
                                village = village_map[vill_code],
                                village_code=record['village_code_census'],
                                land_type=record['land_type'],
                                khata_number=record['khata_number'],
                                khasra_no=record['khasra_no'],
                                area=float(record['area']),
                                name=record['name']
                            )
                            #    print(record['khasra_no'])
                                db.session.add(new_record)

                # Commit the changes
                db.session.commit()
                

@app.route('/', methods=['GET', 'POST'])
def search_form():
    global greeting_flag
    print(greeting_flag)
    if  not greeting_flag:
        # with open('config.json') as config_file:
        #     config = json.load(config_file)
        # data = get_text_message_input(config['RECIPIENT_WAID']
        #                             , 'Welcome to the Land Portal Demo App for Python!')
        # await send_message(data)
        messenger.send_message('Welcome to the Land Portal Demo App for Python!',config['RECIPIENT_WAID'])
        greeting_flag = bool(True)
    print(greeting_flag)
    form = LandSearchForm()
    form.district.choices = [('', 'Any')] + [(district, district) for district in district_tehsil_village_mapping.keys()]

    # Populate tehsil choices based on the selected district
    selected_district = form.district.data
    tehsils = []
    if selected_district in district_tehsil_village_mapping:
        tehsils = district_tehsil_village_mapping[selected_district].keys()
        #tehsils = list(district_tehsil_village_mapping[selected_district].keys())
    form.tehsil.choices = [('', 'Any')] + [(tehsil, tehsil) for tehsil in tehsils]

    # Populate village choices based on the selected tehsil
    selected_tehsil = form.tehsil.data
    villages = {}
    if selected_district in district_tehsil_village_mapping and selected_tehsil in district_tehsil_village_mapping[selected_district]:
        villages = district_tehsil_village_mapping[selected_district][selected_tehsil]

    form.village.choices = [('', 'Any')] + [(village_code, village) for village_code, village in villages.items()]

    form.land_type.choices = list(land_types.items())

    if request.method == 'POST':
        if form.validate_on_submit():
            selected_district = form.district.data
            selected_tehsil = '' if form.tehsil.data is None else form.tehsil.data
            print(selected_tehsil)
            selected_village = None
            # Check if the selected village is in the village mapping
            if form.village.data:
                selected_village = district_tehsil_village_mapping[selected_district][selected_tehsil][form.village.data]

            # if selected_district in district_tehsil_village_mapping and selected_tehsil in district_tehsil_village_mapping[selected_district]:
            #     selected_village = district_tehsil_village_mapping[selected_district][selected_tehsil][form.village.data]
            # else:
            #     selected_village = None
            # #selected_village = district_tehsil_village_mapping[selected_district][selected_tehsil][form.village.data]
            selected_land_type = form.land_type.data
            minimum_area = max(form.area.data, 0.01)
            parameters = {
                "District": selected_district if selected_district != '' else "Any", 
                "Tehsil": selected_tehsil if selected_tehsil != '' else "Any",
                "Village": selected_village if selected_village else "Any",
                "Land Type": selected_land_type if selected_land_type != '' else "Any",
                "Minimum Area": str(minimum_area)
            }
            # Implement your search function here
            # Filter the LandRecord based on the form inputs
            # and_ not needed because in sqlAlchemy filter method they're implictly combined with and_
            # Define the query filter conditions based on the form inputs
            filters = []

            # Nested if conditions to handle hierarchical filtering
            if selected_district != '':
                filters.append(LandRecord.district == selected_district)
                if selected_tehsil != '':
                    filters.append(LandRecord.tehsil == selected_tehsil)
                    if selected_village:
                        filters.append(LandRecord.village_code == selected_village)
            if selected_land_type != '':
                filters.append(LandRecord.land_type == selected_land_type)
            filters.append(LandRecord.area >= minimum_area)
            
            results = LandRecord.query.filter(*filters).all()
            s_results = serialize_land_records(results)
            for index, result in enumerate(results):
                print(index)
                print(result.khata_number)
            
            # Render the 'results.html' template with the search results and parameters
            return render_template('results.html', results=results, s_results = s_results, parameters=parameters, total_count=len(results))
        else:
            print(form.errors)
    return render_template('search_form.html', form=form, district_data=district_tehsil_village_mapping, land_types = land_types)


# Route to generate and download the PDF
@app.route('/download_pdf', methods=['POST'])
def download_pdf():
    # Retrieve the search results and parameters from the form
    s_results = request.form.get('s_results', None)
    # Replace single quotes with double quotes in the JSON string
    s_results = s_results.replace("'", '"')

    parameters = request.form.get('parameters', None)
    parameters = parameters.replace("'", '"')
    parameters_dict = json.loads(parameters)

    results = deserialize_land_records(json.loads(s_results))
    # Render the 'results.html' template with the search results
    rendered_results = render_template('results.html', results=results, s_results = s_results,parameters=parameters_dict, total_count=len(results))

    # Configure PDF options (optional, adjust as needed)
    pdf_options = {
        'page-size': 'Letter',
        'margin-top': '0.75in',
        'margin-right': '0.75in',
        'margin-bottom': '0.75in',
        'margin-left': '0.75in',
    }

    config = pdfkit.configuration(wkhtmltopdf = "C:/Program Files/wkhtmltopdf/bin/wkhtmltopdf.exe")
    # Convert the rendered HTML to PDF using pdfkit
    pdf_file = 'search_results.pdf'
    pdfkit.from_string(rendered_results, pdf_file, options=pdf_options, configuration=config)
            
    # Return the generated PDF file as a response for download
    return send_file(pdf_file, as_attachment=True, download_name='search_results.pdf')

# Route to WhatsApp the PDF
@app.route('/whatsapp_pdf', methods=['POST'])
async def whatsapp_pdf():    
    # Retrieve the search results and parameters from the form
    s_results = request.form.get('s_results', None)
    # Replace single quotes with double quotes in the JSON string
    s_results = s_results.replace("'", '"')
    results = deserialize_land_records(json.loads(s_results))
    
    
    parameters = request.form.get('parameters', None)
    parameters = parameters.replace("'", '"')
    parameters_dict = json.loads(parameters)
    
    link = "C:/Users/praty/Downloads/search_results.pdf"
    try:
        # media_id = upload_media(link, 'application/pdf')
        # print(f"media id is {media_id}")
        # print(f"media_url is {get_media_url(media_id)}")

        media_id = messenger.upload_media(media=link)['id']
        
        messenger.send_document(
        document= media_id,
        recipient_id=config['RECIPIENT_WAID'],
        link = False,
        caption = "Search Results"    
    )
        # data = get_templated_message_input(config['RECIPIENT_WAID'], parameters_dict, get_media_url(media_id), "pdf")
        # await send_message(data)

    except:
        print("File not found.")

    finally:
        return render_template('results.html', results=results, s_results = s_results,parameters=parameters_dict, total_count=len(results))

# Route to generate and download the Excel
@app.route('/download_excel', methods=['POST'])
def download_excel():
    # Retrieve the search results and parameters from the form
    s_results = request.form.get('s_results', None)
    # Replace single quotes with double quotes in the JSON string
    s_results = s_results.replace("'", '"')

    parameters = request.form.get('parameters', None)
    parameters = parameters.replace("'", '"')
    parameters_dict = json.loads(parameters)

    results = json.loads(s_results)  # Convert JSON string to list of dictionaries
    # Render the 'results.html' template with the search results
    #rendered_results = render_template('results.html', results=results, s_results = s_results,parameters=parameters_dict, total_count=len(results))

    wb = openpyxl.Workbook()
    wb.create_sheet(index = 0, title = 'Results')
    #wb.remove_sheet(wb.get_sheet_by_name('Sheet1'))
    sheet = wb.get_sheet_by_name('Results')
    headers = [
        'S.No',
        'district',
        'tehsil',
        'village',
        'village_code',
        'land_type',
        'khata_number',
        'fasli_year',
        'khasra_no',
        'area',
        'name'
    ]
    cols = len(headers)
    for colNum in range(cols):
        sheet.cell(row = 1, column = colNum + 1).value = headers[colNum]
    for rowNum, result in enumerate(results, start = 2):
        sheet.cell(row = rowNum, column = 1).value = rowNum - 1
        for colNum in range(1, cols):
            header = headers[colNum]
            sheet.cell(row = rowNum, column = colNum + 1).value = result[header]
    link = 'C:/Users/praty/OneDrive/Desktop/SDM/Codes/Projects/LandPortal/search_results.xlsx'
    wb.save(link)   
    return send_file(link, as_attachment=True, download_name='search_results.xlsx')     
    

# Route to WhatsApp the PDF
@app.route('/whatsapp_excel', methods=['POST'])
async def whatsapp_excel():    
    # Retrieve the search results and parameters from the form
    s_results = request.form.get('s_results', None)
    # Replace single quotes with double quotes in the JSON string
    s_results = s_results.replace("'", '"')
    results = deserialize_land_records(json.loads(s_results))
    
    
    parameters = request.form.get('parameters', None)
    parameters = parameters.replace("'", '"')
    parameters_dict = json.loads(parameters)
    
    link = "C:/Users/praty/Downloads/search_results.xlsx"
    try:
        # media_id = upload_media(link, 'application/vnd.ms-excel')
        # print(f"media id is {media_id}")
        # print(f"media_url is {get_media_url(media_id)}")
        media_id = messenger.upload_media(media=link)['id']
        messenger.send_document(
        document= media_id,
        recipient_id=config['RECIPIENT_WAID'],
        link = False,
        caption = "Search Results"    
    )
        # data = get_templated_message_input(config['RECIPIENT_WAID'], parameters_dict, get_media_url(media_id), "xlsx")
        # await send_message(data)

    except:
        print("File not found.")

    finally:
        return render_template('results.html', results=results, s_results = s_results,parameters=parameters_dict, total_count=len(results))



if __name__ == '__main__':
    app.run(debug=True)
